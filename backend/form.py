from sqlalchemy import Column, Integer, String, ForeignKey, DateTime, func, Boolean, text, or_
from sqlalchemy.orm import relationship, Session, joinedload
from database import Base
from pydantic import BaseModel
from fastapi import FastAPI, HTTPException, Depends, APIRouter, UploadFile, File, Form, Request, Query, Header
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder
from fastapi.datastructures import UploadFile
from io import BytesIO
from statistics import mean
from openpyxl import Workbook
from form_db_circl_fix import get_db
from form_cur_user_circl_fix import get_current_user, get_current_user_optional
from schemas import QuestionCreate, QuestionCreateMultipart, OptionWithImage
from models import QuestionModel, OptionModel, UserModel,AnswerModel, FormModel, CollaboratorModel
from schemas import FormOut, FormCreate, AnswerSubmission, FormUpdate, QuestionUpdate, CollaboratorIn, CollaboratorOut, FormWithMeta, CollaboratorUpdate, CollaboratorRemove
from typing import Optional, List, Tuple
import json, os, uuid, shutil
from uuid import uuid4
import mimetypes
from cloudinary_utils import upload_to_cloudinary

router = APIRouter()

def is_superadmin(user: Optional[UserModel]) -> bool:
    return bool(user and getattr(user, "is_superadmin", False))

def _actor_role_for_form(
    db: Session,
    form: FormModel,
    current_user: Optional[UserModel],
    as_user: Optional[int],
    x_impersonate_user: Optional[int],
) -> Tuple[UserModel, str, bool]:
    if is_superadmin(current_user):
        if as_user is not None or x_impersonate_user is not None:
            _, _ = resolve_actor_for_impersonation(db, current_user, as_user, x_impersonate_user)

        return current_user, "superadmin", (as_user is not None or x_impersonate_user is not None)

    
    actor = current_user
    role = "viewer"
    if actor:
        if actor.id == form.owner_id:
            role = "owner"
        else:
            collab = db.query(CollaboratorModel).filter(
                CollaboratorModel.form_id == form.id,
                CollaboratorModel.user_id == actor.id,
            ).first()
            if collab:
                role = "editor" if (collab.role or "").lower() == "editor" else "viewer"

    return actor, role, False


def _ensure_can_edit_form(
    db: Session,
    form: FormModel,
    current_user: Optional[UserModel],
    as_user: Optional[int],
    x_impersonate_user: Optional[int],
    *,
    owner_only: bool = False,
):
    
    if is_superadmin(current_user):
        return current_user, "superadmin"

    actor, role, _ = _actor_role_for_form(db, form, current_user, as_user, x_impersonate_user)

    if owner_only:
        if role != "owner":
            raise HTTPException(status_code=403, detail="Only owner can perform this action.")
    else:
        if role not in ("owner", "editor"):
            raise HTTPException(status_code=403, detail="Not authorized to edit this form.")
    return actor, role


def resolve_actor_for_impersonation(
    db: Session,
    current_user: UserModel,
    as_user: Optional[int] = None,
    x_impersonate_user: Optional[int] = None,
) -> Tuple[UserModel, bool]:
    if as_user is None and x_impersonate_user is None:
        return current_user, False

    if as_user is not None and x_impersonate_user is not None and as_user != x_impersonate_user:
        raise HTTPException(status_code=400, detail="Conflicting as_user and X-Impersonate-User")

    if not is_superadmin(current_user):
        raise HTTPException(status_code=403, detail="as_user allowed only for superadmin")

    target_id = as_user if as_user is not None else x_impersonate_user
    target = db.query(UserModel).filter(UserModel.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Impersonated user not found")

    return target, True

def _collab_role(db: Session, form_id: int, user_id: int) -> Optional[str]:
    row = db.query(CollaboratorModel.role).filter(
        CollaboratorModel.form_id == form_id,
        CollaboratorModel.user_id == user_id
    ).first()
    return row[0] if row else None

def _role_and_caps(db: Session, form: FormModel, user: Optional[UserModel]):
    role = None
    can_edit = False
    if user:
        if form.owner_id == user.id:
            role, can_edit = "owner", True
        else:
            r = _collab_role(db, form.id, user.id)
            if r:
                role = r                     
                can_edit = (r == "editor")
    return role, {"can_edit": can_edit}

def _can_edit(form: FormModel, user: Optional[UserModel], db: Session) -> bool:
    if not user:
        return False
    if form.owner_id == user.id:
        return True
    return _collab_role(db, form.id, user.id) == "editor"

def _can_view_form(db: Session, form: FormModel, user: Optional[UserModel]) -> bool:
    if form.is_public:
        return True
    if is_superadmin(user):
        return True
    if not user:
        return False
    if user.id == form.owner_id:
        return True
    collab = db.query(CollaboratorModel).filter(
        CollaboratorModel.form_id == form.id,
        CollaboratorModel.user_id == user.id,
    ).first()
    return collab is not None

def _can_view_results(form: FormModel, user: Optional[UserModel], db: Session) -> bool:
    if is_superadmin(user):
        return True
    if form.is_public:
        return True
    if not user:
        return False
    if user.id == form.owner_id:
        return True
    collab = db.query(CollaboratorModel).filter(
        CollaboratorModel.form_id == form.id,
        CollaboratorModel.user_id == user.id,
    ).first()
    return collab is not None


def _can_view_collaborators(db: Session, form: FormModel, user: Optional[UserModel]) -> bool:
    if is_superadmin(user):
        return True
    if not user:
        return False
    if user.id == form.owner_id:
        return True
    collab = db.query(CollaboratorModel).filter(
        CollaboratorModel.form_id == form.id,
        CollaboratorModel.user_id == user.id
    ).first()
    return bool(collab and (collab.role or "").lower() == "editor")

def _ensure_can_manage_collaborators(db: Session, form: FormModel, user: Optional[UserModel]):
    if is_superadmin(user):
        return
    if not user or user.id != form.owner_id:
        raise HTTPException(status_code=403, detail="Only owner (or superadmin) can manage collaborators.")
    
def _fmt_num(x: float) -> str:
    # lepa reprezentacija: 2 umesto 2.0
    return str(int(x)) if float(x).is_integer() else str(x)
    
def _gen_scale(start: float, end: float, step: float) -> list[str]:
    if step == 0:
        raise HTTPException(400, "numeric_scale.step must not be 0")
    values = []
    v = start
    
    if (end - start) * step < 0:
        raise HTTPException(400, "numeric_scale has inconsistent direction")
    
    eps = abs(step) * 1e-9
    if step > 0:
        while v <= end + eps:
            values.append(_fmt_num(v))
            v += step
    else:
        while v >= end - eps:
            values.append(_fmt_num(v))
            v += step
    return values

def _normalize_numeric_choice_options(payload: "QuestionCreate") -> list[dict]:
    """
    Vrati listu opcija u formatu [{'text': '...'}, ...] za numeric_choice.
    Prihvatamo: options (stringovi), numeric_values (brojevi) ili numeric_scale.
    """
    if payload.options:
        
        out = []
        for o in payload.options:
            t = str(o.get("text", ""))
            try:
                float(t)
            except ValueError:
                raise HTTPException(400, f"numeric_choice option '{t}' is not a number")
            out.append({"text": t})
        return out

    if payload.numeric_values:
        return [{"text": _fmt_num(float(x))} for x in payload.numeric_values]

    if payload.numeric_scale:
        s = payload.numeric_scale
        return [{"text": t} for t in _gen_scale(float(s.start), float(s.end), float(s.step))]

    raise HTTPException(400, "numeric_choice requires options OR numeric_values OR numeric_scale")

def _build_analytics_for_form(db: Session, form: FormModel):
    result = []
    for q in form.questions:
        q_summary = {"question_id": q.id, "text": q.text, "type": q.type}
        answers = db.query(AnswerModel).filter(AnswerModel.question_id == q.id).all()
        total = len(answers)

        opt_by_id = {o.id: o for o in q.options}
        opt_by_text = { (o.text or "").strip().lower(): o for o in q.options }

        if q.type in ("radio","single_choice"):
            counts = {o.id: 0 for o in q.options}
            other = 0
            for a in answers:
                val = (a.value or "").strip()
                hit = False
                try:
                    oid = int(val)
                    if oid in counts:
                        counts[oid] += 1
                        hit = True
                except:
                    pass
                if not hit and val:
                    o = opt_by_text.get(val.lower())
                    if o:
                        counts[o.id] += 1
                        hit = True
                if not hit:
                    other += 1
            items = []
            for oid, cnt in counts.items():
                pct = (cnt/total*100.0) if total else 0.0
                items.append({
                    "option_id": oid,
                    "option_text": opt_by_id[oid].text,
                    "count": cnt,
                    "percent": round(pct, 2)
                })
            if other:
                items.append({"option_id": None, "option_text": "_other", "count": other, "percent": round(other/total*100.0, 2)})
            q_summary["distribution"] = items
            q_summary["total_answers"] = total

        elif q.type in ("checkbox","multiple_choice"):
            counts = {o.id: 0 for o in q.options}
            other = 0
            total_selections = 0
            for a in answers:
                tokens = [t.strip() for t in (a.value or "").split(",") if t.strip()]
                for tok in tokens:
                    total_selections += 1
                    hit = False
                    try:
                        oid = int(tok)
                        if oid in counts:
                            counts[oid] += 1
                            hit = True
                    except:
                        pass
                    if not hit:
                        o = opt_by_text.get(tok.lower())
                        if o:
                            counts[o.id] += 1
                            hit = True
                    if not hit:
                        other += 1
            items = []
            denom = total_selections if total_selections else 1
            for oid, cnt in counts.items():
                pct = (cnt/denom*100.0)
                items.append({
                    "option_id": oid,
                    "option_text": opt_by_id[oid].text,
                    "count": cnt,
                    "percent": round(pct, 2)
                })
            if other:
                items.append({"option_id": None, "option_text": "_other", "count": other, "percent": round(other/denom*100.0, 2)})
            q_summary["distribution"] = items
            q_summary["total_answers"] = total
            q_summary["total_selections"] = total_selections

        elif q.type in ("number","numeric"):
            nums = []
            freq = {}
            for a in answers:
                try:
                    x = float(a.value)
                    nums.append(x)
                    freq[a.value] = freq.get(a.value, 0) + 1
                except:
                    pass
            q_summary["count"] = len(nums)
            if nums:
                q_summary["min"] = min(nums)
                q_summary["avg"] = round(mean(nums), 4)
                q_summary["max"] = max(nums)
            q_summary["histogram"] = [{"value": k, "count": v} for k, v in sorted(freq.items(), key=lambda kv: float(kv[0]))]

        elif q.type in ("date","time"):
            freq = {}
            for a in answers:
                v = (a.value or "").strip()
                if v:
                    freq[v] = freq.get(v, 0) + 1
            q_summary["count"] = total
            q_summary["by_value"] = [{"value": k, "count": v} for k, v in sorted(freq.items())]

        else:
            q_summary["count"] = total

        result.append(q_summary)
    return result




@router.post("/api/forms")
def create_form(data: FormCreate, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    new_form = FormModel(
        name=data.name,
        description=data.description,
        is_public=data.is_public,
        owner_id = current_user.id 


    )
    db.add(new_form)
    db.commit()
    db.refresh(new_form)
    return {"id": new_form.id, "message": "Form created successfully"}


@router.get("/api/forms/owned", response_model=List[FormWithMeta])
def get_owned_forms(
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    forms = db.query(FormModel).filter(FormModel.owner_id == current_user.id).all()
    out = []
    for f in forms:
        role, caps = _role_and_caps(db, f, current_user) 
        out.append({
            "id": f.id,
            "name": f.name,
            "description": f.description,
            "is_public": f.is_public,
            "questions": None,
            "my_role": role,
            "capabilities": caps
        })
    return out


@router.get("/api/forms/public", response_model=List[FormOut])
def get_public_forms(db: Session = Depends(get_db)):
    forms = db.query(FormModel).filter(FormModel.is_public == True).all()
    return forms



@router.get("/api/forms/mine", response_model=List[FormWithMeta])
def get_my_forms(
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    owner_q = db.query(FormModel).filter(FormModel.owner_id == current_user.id)

    collab_form_ids = db.query(CollaboratorModel.form_id).filter(
        CollaboratorModel.user_id == current_user.id
    )
    collab_q = db.query(FormModel).filter(FormModel.id.in_(collab_form_ids))

    forms = owner_q.union_all(collab_q).all()

    out: List[dict] = []
    for f in forms:
        role, caps = _role_and_caps(db, f, current_user)
        out.append({
            "id": f.id,
            "name": f.name,
            "description": f.description,
            "is_public": f.is_public,
            "questions": None,      
            "my_role": role,
            "capabilities": caps
        })
    return out

@router.get("/api/users/{user_id}/forms")
def list_user_forms(
    user_id: int,
    include_collab: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    if not (current_user.is_superadmin or current_user.id == user_id):
        raise HTTPException(status_code=403, detail="Nemaš dozvolu za pristup formama ovog korisnika.")

    forms = db.query(FormModel).filter(FormModel.owner_id == user_id).all()

    if include_collab:
        collab_ids = db.query(CollaboratorModel.form_id).filter(CollaboratorModel.user_id == user_id)
        collab_forms = db.query(FormModel).filter(FormModel.id.in_(collab_ids)).all()
        
        by_id = {f.id: f for f in forms}
        for f in collab_forms:
            by_id.setdefault(f.id, f)
        forms = list(by_id.values())

    return [
        {
            "id": f.id,
            "name": f.name,
            "description": f.description,
            "is_public": f.is_public,
        }
        for f in forms
    ]

@router.post("/api/forms/{form_id}/questions")
def add_question(
    form_id: int,
    request: Request,
    text: str = Form(...),
    type: str = Form(...),
    is_required: bool = Form(...),
    order: Optional[int] = Form(None),
    max_choices: Optional[int] = Form(None),
    options: Optional[str] = Form(None),            
    numeric_scale: Optional[str] = Form(None),      
    image: Optional[UploadFile] = File(None),
    option_images: List[UploadFile] = File([]),
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

   
    actor, _ = resolve_actor_for_impersonation(db, current_user, as_user, x_impersonate_user)
    _ensure_can_edit_form(db, form, current_user, as_user, x_impersonate_user, owner_only=False)

    
    type_alias = {"radio": "single_choice", "checkbox": "multiple_choice"}
    qtype = type_alias.get(type, type)

    allowed_types = {
        "single_choice", "multiple_choice",
        "numeric_choice", "short_text", "long_text",
        "date", "datetime"
    }
    if qtype not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Unsupported question type: {qtype}")

    
    effective_max_choices = max_choices if qtype == "multiple_choice" else None

    
    image_url = None
    if image:
        ext = os.path.splitext(image.filename)[1]
        filename = f"{uuid4().hex}{ext}"
        filepath = os.path.join("temp", filename)
        os.makedirs("temp", exist_ok=True)
        with open(filepath, "wb") as f:
            f.write(image.file.read())
        image_url = upload_to_cloudinary(filepath, filename)
        os.remove(filepath)

    
    new_question = QuestionModel(
        form_id=form_id,
        text=text,
        type=qtype,
        is_required=is_required,
        order=order,
        max_choices=effective_max_choices,
        image_url=image_url
    )
    db.add(new_question)
    db.commit()
    db.refresh(new_question)

    
    def _gen_numeric_series(start: float, end: float, step: float) -> list[float]:
        if step == 0:
            raise HTTPException(status_code=400, detail="numeric_scale.step cannot be 0")
        vals = []
        v = float(start)
        step = float(step)
        
        if start <= end and step < 0:
            step = abs(step)
        if start > end and step > 0:
            step = -step
        
        for _ in range(10000):
            if (step > 0 and v > end) or (step < 0 and v < end):
                break
            
            vals.append(float(f"{v:.10g}"))
            v += step
        return vals

    if qtype in ("single_choice", "multiple_choice", "numeric_choice"):
        parsed_options = None

        if options:
            try:
                parsed_options = json.loads(options)  
                if not isinstance(parsed_options, list):
                    raise ValueError
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid options JSON format")

        elif qtype == "numeric_choice" and numeric_scale:
            
            try:
                ns = json.loads(numeric_scale)  
                start = float(ns["start"]); end = float(ns["end"]); step = float(ns["step"])
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid numeric_scale JSON format")
            series = _gen_numeric_series(start, end, step)
            parsed_options = [{"text": str(v)} for v in series]

        
        if qtype == "numeric_choice":
            if not parsed_options:
                raise HTTPException(status_code=400, detail="numeric_choice requires options or numeric_scale")
            
            for opt in parsed_options:
                try:
                    _ = float(str(opt.get("text", "")).strip())
                except Exception:
                    raise HTTPException(
                        status_code=400,
                        detail="numeric_choice options must be numeric"
                    )

        
        if parsed_options:
            for idx, opt in enumerate(parsed_options):
                opt_image_url = None
                
                if qtype in ("single_choice", "multiple_choice") and idx < len(option_images):
                    img_file = option_images[idx]
                    if img_file and img_file.filename:
                        ext = os.path.splitext(img_file.filename)[1]
                        fname = f"{uuid4().hex}{ext}"
                        fpath = os.path.join("temp", fname)
                        os.makedirs("temp", exist_ok=True)
                        with open(fpath, "wb") as f:
                            f.write(img_file.file.read())
                        opt_image_url = upload_to_cloudinary(fpath, fname)
                        os.remove(fpath)

                db.add(OptionModel(
                    text=str(opt["text"]),
                    question_id=new_question.id,
                    image_url=opt_image_url
                ))
            db.commit()

    return {"id": new_question.id, "message": "Question created"}




@router.get("/api/forms/{form_id}", response_model=FormOut)
def get_form_by_id(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[UserModel] = Depends(get_current_user_optional),
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
):
    form = (
        db.query(FormModel)
        .options(joinedload(FormModel.questions).joinedload(QuestionModel.options))
        .filter(FormModel.id == form_id)
        .first()
    )
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    if not _can_view_form(db, form, current_user):
        raise HTTPException(status_code=403, detail="Access denied to private form")

    if is_superadmin(current_user):
        role = "superadmin"
        can_edit = True
    else:
        if current_user and current_user.id == form.owner_id:
            role = "owner"
            can_edit = True
        else:
            collab = (
                db.query(CollaboratorModel)
                .filter(
                    CollaboratorModel.form_id == form_id,
                    CollaboratorModel.user_id == (current_user.id if current_user else -1),
                )
                .first()
            )
            if collab:
                role = "editor" if (collab.role or "").lower() == "editor" else "viewer"
                can_edit = (role == "editor")
            else:
                role = "viewer"
                can_edit = False

    converted_questions = []
    for q in form.questions:
        converted_options = [
            {"id": opt.id, "text": opt.text, "image_url": opt.image_url or None}
            for opt in q.options
        ]
        converted_questions.append(
            {
                "id": q.id,
                "text": q.text,
                "type": q.type,
                "is_required": q.is_required,
                "order": q.order,
                "max_choices": q.max_choices,
                "image_url": q.image_url or None,
                "options": converted_options,
            }
        )

    return {
        "id": form.id,
        "name": form.name,
        "description": form.description,
        "is_public": form.is_public,
        "is_locked": form.is_locked,
        "questions": converted_questions,
        "my_role": role,
        "capabilities": {"can_edit": can_edit},
    }

@router.delete("/api/forms/{form_id}", status_code=200)
def delete_form(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[UserModel] = Depends(get_current_user_optional),
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma nije pronađena.")
    
    if not is_superadmin(current_user) and (not current_user or current_user.id != form.owner_id):
        raise HTTPException(status_code=403, detail="Samo vlasnik ili superadmin može da obriše formu.")

    db.query(CollaboratorModel).filter(CollaboratorModel.form_id == form_id).delete(synchronize_session=False)

    db.execute(
        text("DELETE FROM answers WHERE question_id IN (SELECT id FROM questions WHERE form_id=:fid)"),
        {"fid": form_id},
    )

    db.delete(form)
    db.commit()
    return {"ok": True}
@router.post("/api/forms/{form_id}/answers")
def submit_answers(
    form_id: int,
    submission: AnswerSubmission,
    db: Session = Depends(get_db),
    current_user: Optional[UserModel] = Depends(get_current_user_optional)
):
    
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")
    if form.is_locked and not (current_user and getattr(current_user, "is_superadmin", False)):
        raise HTTPException(status_code=403, detail="Forma je zaključana za odgovore.")
    if not form.is_public and current_user is None:
        raise HTTPException(status_code=403, detail="Morate biti prijavljeni da biste popunili ovu formu.")

    
    qrows = (
        db.query(QuestionModel)
        .filter(QuestionModel.form_id == form_id)
        .order_by(QuestionModel.order)
        .all()
    )
    questions_by_id = {q.id: q for q in qrows}

    
    opts_by_qid: dict[int, list[OptionModel]] = {}
    if qrows:
        qids = [q.id for q in qrows]
        all_opts = db.query(OptionModel).filter(OptionModel.question_id.in_(qids)).all()
        for o in all_opts:
            opts_by_qid.setdefault(o.question_id, []).append(o)

    
    sub_by_qid = {a.question_id: a for a in submission.answers}

    
    for q in qrows:
        if q.is_required:
            a = sub_by_qid.get(q.id)
            if a is None:
                raise HTTPException(status_code=400, detail=f"Pitanje '{q.text}' je obavezno.")
            if a.answer is None or (isinstance(a.answer, str) and a.answer.strip() == ""):
                
                if not (q.type in ("multiple_choice", "checkbox") and isinstance(a.answer, list) and len(a.answer) > 0):
                    raise HTTPException(status_code=400, detail=f"Pitanje '{q.text}' je obavezno.")

    def _norm_type(t: str) -> str:
        alias = {"radio": "single_choice", "checkbox": "multiple_choice"}
        return alias.get(t, t)

    def _stringy(x) -> str:
        return "" if x is None else str(x)

    
    def _parse_id_like(x) -> Optional[int]:
        try:
            return int(str(x).strip())
        except Exception:
            return None

    def _validate_single(q: QuestionModel, raw) -> str:
        """Prihvata option id ili option text; vraća string koji snimamo (ID kao string)."""
        options = opts_by_qid.get(q.id, [])
        if not options:
            raise HTTPException(status_code=400, detail=f"Pitanje '{q.text}' nema opcije.")
        
        oid = _parse_id_like(raw)
        if oid is not None and any(o.id == oid for o in options):
            return str(oid)
        
        sval = str(raw)
        match = next((o for o in options if o.text == sval), None)
        if match:
            return str(match.id)
        raise HTTPException(status_code=400, detail=f"Nevažeća opcija za pitanje '{q.text}'.")

    def _validate_multi(q: QuestionModel, raw) -> str:
        """Prihvata listu ID-jeva ili tekstova; vraća CSV string sa ID-jevima."""
        options = opts_by_qid.get(q.id, [])
        if not options:
            raise HTTPException(status_code=400, detail=f"Pitanje '{q.text}' nema opcije.")
        
        if isinstance(raw, str):
            cand = [p.strip() for p in raw.split(",") if p.strip()]
        elif isinstance(raw, list):
            cand = raw
        else:
            raise HTTPException(status_code=400, detail=f"Multiple choice za '{q.text}' očekuje listu ili CSV string.")

        ids: list[int] = []
        for it in cand:
            oid = _parse_id_like(it)
            if oid is not None and any(o.id == oid for o in options):
                ids.append(oid)
                continue
        
            sval = str(it)
            match = next((o for o in options if o.text == sval), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"Nevažeća opcija '{sval}' za '{q.text}'.")
            ids.append(match.id)

        
        if q.max_choices is not None and len(ids) > int(q.max_choices):
            raise HTTPException(
                status_code=400,
                detail=f"Dozvoljeno je najviše {q.max_choices} izbora za '{q.text}'."
            )
        
        return ",".join(str(i) for i in ids)

    def _validate_numeric_choice(q: QuestionModel, raw) -> str:
        """Prihvata ID opcije ili numeričku vrednost koja postoji među opcijama (po tekstu)."""
        options = opts_by_qid.get(q.id, [])
        if not options:
            raise HTTPException(status_code=400, detail=f"Pitanje '{q.text}' nema opcije.")
        
        oid = _parse_id_like(raw)
        if oid is not None and any(o.id == oid for o in options):
            return str(oid)
        
        try:
            f = float(str(raw).strip())
        except Exception:
            raise HTTPException(status_code=400, detail=f"Odgovor za '{q.text}' mora biti broj.")
        
        for o in options:
            try:
                if float(o.text) == f:
                    return str(o.id)
            except Exception:
                continue
        raise HTTPException(status_code=400, detail=f"Odgovor nije u dozvoljenom opsegu za '{q.text}'.")

    def _validate_date(raw) -> str:
        from datetime import date
        try:
            return date.fromisoformat(str(raw)).isoformat()
        except Exception:
            raise HTTPException(status_code=400, detail="Neispravan datum (očekivan ISO format YYYY-MM-DD).")

    def _validate_datetime(raw) -> str:
        from datetime import datetime
        s = str(raw).replace("Z", "+00:00")
        try:
            
            dt = datetime.fromisoformat(s)
            return dt.replace(microsecond=0).isoformat()
        except Exception:
            raise HTTPException(status_code=400, detail="Neispravan datum/vreme (ISO 8601).")

    
    for item in submission.answers:
        q = questions_by_id.get(item.question_id)
        if not q:
            raise HTTPException(status_code=400, detail=f"Nevažeće pitanje ID: {item.question_id}")

        qtype = _norm_type(q.type)

        if qtype == "single_choice":
            stored_val = _validate_single(q, item.answer)
        elif qtype == "multiple_choice":
            stored_val = _validate_multi(q, item.answer)
        elif qtype == "numeric_choice":
            stored_val = _validate_numeric_choice(q, item.answer)
        elif qtype in ("short_text", "long_text"):
            stored_val = _stringy(item.answer)
        elif qtype == "date":
            stored_val = _validate_date(item.answer)
        elif qtype == "datetime":
            stored_val = _validate_datetime(item.answer)
        else:
            
            stored_val = _stringy(item.answer)

        db.add(AnswerModel(
            question_id=item.question_id,
            user_id=(current_user.id if current_user else None),
            value=stored_val,
        ))

    db.commit()
    return {"message": "Odgovori su uspešno sačuvani."}

@router.get("/api/forms/{form_id}/answers")
def get_answers_for_form(
    form_id: int,
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")

    actor, _ = resolve_actor_for_impersonation(db, current_user, as_user, x_impersonate_user)
    if not _can_view_results(form, actor, db):
        raise HTTPException(403, "Nemate pravo da vidite rezultate.")

    answers = db.execute(
        text("""
            SELECT q.text AS question_text,
                   a.value AS answer_value,
                   u.email AS user_email
            FROM answers a
            JOIN questions q ON q.id = a.question_id
            LEFT JOIN users u ON u.id = a.user_id
            WHERE q.form_id = :fid
        """),
        {"fid": form_id}
    ).fetchall()

    results = []
    for row in answers:
        results.append({
            "pitanje": row.question_text,
            "odgovor": row.answer_value,
            "korisnik": row.user_email if row.user_email else "anonymo"
        })

    return JSONResponse(content=results)


@router.put("/api/forms/{form_id}")
def update_form(
    form_id: int,
    data: FormUpdate,
    db: Session = Depends(get_db),
    current_user: Optional[UserModel] = Depends(get_current_user_optional),
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")

    _ensure_can_edit_form(db, form, current_user, as_user, x_impersonate_user, owner_only=False)

    if data.name is not None:
        form.name = data.name
    if data.description is not None:
        form.description = data.description
    if data.is_public is not None:
        form.is_public = data.is_public
    if data.is_locked is not None:                 
        form.is_locked = data.is_locked

    db.commit()
    db.refresh(form)
    return {
        "id": form.id,
        "name": form.name,
        "description": form.description,
        "is_public": form.is_public,
        "is_locked": form.is_locked,
        "owner_id": form.owner_id,
    }


@router.put("/api/forms/{form_id}/questions/{question_id}")
def update_question(
    form_id: int,
    question_id: int,
    updated_data: QuestionUpdate,                 
    db: Session = Depends(get_db),
    current_user: Optional[UserModel] = Depends(get_current_user_optional),
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
):
    
    question = (
        db.query(QuestionModel)
        .filter(QuestionModel.id == question_id, QuestionModel.form_id == form_id)
        .first()
    )
    if not question:
        raise HTTPException(status_code=404, detail="Pitanje nije pronađeno")

    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")

    
    _ensure_can_edit_form(db, form, current_user, as_user, x_impersonate_user, owner_only=False)

    
    alias = {"radio": "single_choice", "checkbox": "multiple_choice"}
    if updated_data.type is not None:
        qtype = alias.get(updated_data.type, updated_data.type)
        allowed = {
            "single_choice", "multiple_choice",
            "numeric_choice", "short_text", "long_text",
            "date", "datetime"
        }
        if qtype not in allowed:
            raise HTTPException(status_code=400, detail=f"Unsupported question type: {qtype}")
        question.type = qtype
    else:
        qtype = alias.get(question.type, question.type)

    if updated_data.text is not None:
        question.text = updated_data.text

    
    if qtype == "multiple_choice":
        if updated_data.max_choices is not None:
            question.max_choices = updated_data.max_choices
    else:
        question.max_choices = None

    
    def _gen_numeric_series(start: float, end: float, step: float) -> list[float]:
        if step == 0:
            raise HTTPException(status_code=400, detail="numeric_scale.step cannot be 0")
        vals = []
        v = float(start)
        step = float(step)
        if start <= end and step < 0:
            step = abs(step)
        if start > end and step > 0:
            step = -step
        for _ in range(10000):
            if (step > 0 and v > end) or (step < 0 and v < end):
                break
            vals.append(float(f"{v:.10g}"))
            v += step
        return vals

    
    has_options = updated_data.options is not None
    has_scale = getattr(updated_data, "numeric_scale", None) is not None

    if has_options or has_scale:
        
        db.query(OptionModel).filter(OptionModel.question_id == question_id).delete()

        new_texts: list[str] = []
        if has_scale:
            ns = updated_data.numeric_scale
            series = _gen_numeric_series(float(ns.start), float(ns.end), float(ns.step))
            new_texts = [str(v) for v in series]
        elif has_options:
            new_texts = [str(opt.text) for opt in updated_data.options]

        
        if qtype == "numeric_choice":
            if not new_texts:
                raise HTTPException(status_code=400, detail="numeric_choice requires options or numeric_scale")
            for t in new_texts:
                try:
                    float(t.strip())
                except Exception:
                    raise HTTPException(status_code=400, detail="numeric_choice options must be numeric")

        for t in new_texts:
            db.add(OptionModel(text=t, question_id=question_id))

    db.commit()
    db.refresh(question)

    return {
        "id": question.id,
        "text": question.text,
        "type": question.type,
        "form_id": question.form_id,
        "max_choices": question.max_choices,
    }

@router.get("/api/forms/{form_id}/collaborators", response_model=List[CollaboratorOut])
def list_collaborators(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma ne postoji.")
    if not _can_view_collaborators(db, form, current_user):
        raise HTTPException(status_code=403, detail="Not allowed to view collaborators.")
    rows = (
        db.query(CollaboratorModel, UserModel)
        .join(UserModel, UserModel.id == CollaboratorModel.user_id)
        .filter(CollaboratorModel.form_id == form_id)
        .all()
    )
    return [
        CollaboratorOut(user_id=u.id, username=u.username, email=u.email, role=c.role)
        for c, u in rows
    ]

@router.put("/api/forms/{form_id}/collaborators")
def update_collaborator_role(
    form_id: int,
    data: CollaboratorUpdate,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")

    if not (is_superadmin(current_user) or form.owner_id == current_user.id):
        raise HTTPException(status_code=403, detail="Nemate pravo izmene kolaboratora.")

    coll = (
        db.query(CollaboratorModel)
        .filter(
            CollaboratorModel.form_id == form_id,
            CollaboratorModel.user_id == data.user_id,
        )
        .first()
    )
    if not coll:
        raise HTTPException(status_code=404, detail="Kolaborator ne postoji na ovoj formi.")

    if data.role not in ("viewer", "editor"):
        raise HTTPException(status_code=400, detail="Nevalidna uloga.")

    coll.role = data.role
    db.commit()
    db.refresh(coll)

    user = db.query(UserModel).filter(UserModel.id == coll.user_id).first()
    return {
        "form_id": form_id,
        "user_id": coll.user_id,
        "role": coll.role,
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
        } if user else None,
    }


@router.post("/api/forms/{form_id}/collaborators", status_code=201)
def add_or_update_collaborator(
    form_id: int,
    payload: CollaboratorIn,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma ne postoji.")
    _ensure_can_manage_collaborators(db, form, current_user)

    if payload.role not in ("viewer", "editor"):
        raise HTTPException(422, "Nedozvoljena uloga.")

    if payload.user_id == form.owner_id:
        raise HTTPException(400, "Vlasnik ne može biti dodat kao kolaborator.")

    target_user = db.query(UserModel).filter(UserModel.id == payload.user_id).first()
    if not target_user:
        raise HTTPException(404, "Korisnik ne postoji.")

    existing = db.query(CollaboratorModel).filter(
        CollaboratorModel.form_id == form_id,
        CollaboratorModel.user_id == payload.user_id
    ).first()

    if existing:
        existing.role = payload.role
    else:
        db.add(CollaboratorModel(form_id=form_id, user_id=payload.user_id, role=payload.role))

    db.commit()
    return {"message": "Kolaborator je sačuvan."}


@router.delete("/api/forms/{form_id}/collaborators", status_code=204)
def remove_collaborator(
    form_id: int,
    data: CollaboratorRemove,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Forma ne postoji.")

    if not (is_superadmin(current_user) or form.owner_id == current_user.id):
        raise HTTPException(status_code=403, detail="Nemate pravo brisanja kolaboratora.")

    coll = (
        db.query(CollaboratorModel)
        .filter(
            CollaboratorModel.form_id == form_id,
            CollaboratorModel.user_id == data.user_id,
        )
        .first()
    )
    if not coll:
        raise HTTPException(status_code=404, detail="Kolaborator ne postoji na ovoj formi.")

    db.delete(coll)
    db.commit()
    return

@router.delete("/api/forms/{form_id}/collaborators/{user_id}", status_code=204)
def remove_collaborator(
    form_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma ne postoji.")
    _ensure_can_manage_collaborators(db, form, current_user)

    deleted = db.query(CollaboratorModel).filter(
        CollaboratorModel.form_id == form_id,
        CollaboratorModel.user_id == user_id
    ).delete()
    db.commit()
    if deleted == 0:
        raise HTTPException(404, "Kolaborator nije pronađen.")
    

@router.get("/api/forms/{form_id}/analytics")
def analytics_for_form(
    form_id: int,
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"), 
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).options(
        joinedload(FormModel.questions).joinedload(QuestionModel.options)
    ).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma ne postoji.")

    actor, _ = resolve_actor_for_impersonation(db, current_user, as_user, x_impersonate_user)

    if not _can_view_results(form, actor, db):
        raise HTTPException(403, "Nemate pravo pregleda rezultata.")

    data = _build_analytics_for_form(db, form)
    return {"form_id": form.id, "analytics": data}


@router.get("/api/forms/{form_id}/export.xlsx")
def export_xlsx(
    form_id: int,
    as_user: Optional[int] = Query(None),
    x_impersonate_user: Optional[int] = Header(None, alias="X-Impersonate-User"),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    form = db.query(FormModel).options(
        joinedload(FormModel.questions).joinedload(QuestionModel.options)
    ).filter(FormModel.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma ne postoji.")

    actor, _ = resolve_actor_for_impersonation(db, current_user, as_user, x_impersonate_user)

    if not _can_view_results(form, actor, db):
        raise HTTPException(403, "Nemate pravo izvoza rezultata.")

    wb = Workbook()
    ws_q = wb.active
    ws_q.title = "Questions"
    ws_q.append(["Question ID", "Text", "Type", "Required", "Order", "Max choices", "Options"])

    for q in form.questions:
        opts = " | ".join([o.text for o in q.options]) if q.options else ""
        ws_q.append([q.id, q.text, q.type, bool(q.is_required), q.order, q.max_choices, opts])

    ws_a = wb.create_sheet("Answers")
    ws_a.append(["Question ID", "Question", "Answer", "User email", "Submitted at"])

    rows = db.execute(
        text("""
            SELECT q.id AS qid, q.text AS qtext, a.value AS answer_value,
                   u.email AS user_email, a.submitted_at
            FROM answers a
            JOIN questions q ON q.id = a.question_id
            LEFT JOIN users u ON u.id = a.user_id
            WHERE q.form_id = :fid
            ORDER BY a.submitted_at ASC
        """),
        {"fid": form_id}
    ).fetchall()

    for r in rows:
        ws_a.append([r.qid, r.qtext, r.answer_value, (r.user_email or "anonimo"), r.submitted_at])

    ws_an = wb.create_sheet("Analytics")
    ws_an.append(["Question ID", "Text", "Type", "Metric", "Value", "Extra"])

    analytics = _build_analytics_for_form(db, form)
    for item in analytics:
        qt, qid, qtype = item["text"], item["question_id"], item["type"]

        if "distribution" in item:
            for d in item["distribution"]:
                ws_an.append([qid, qt, qtype, "option",
                              f'{d.get("option_text")}', f'count={d["count"]}, percent={d["percent"]}%'])
            ws_an.append([qid, qt, qtype, "total_answers", item.get("total_answers", 0), ""])
            if "total_selections" in item:
                ws_an.append([qid, qt, qtype, "total_selections", item["total_selections"], ""])
        elif qtype in ("number","numeric"):
            ws_an.append([qid, qt, qtype, "count", item.get("count", 0), ""])
            if "min" in item:
                ws_an.append([qid, qt, qtype, "min", item["min"], ""])
                ws_an.append([qid, qt, qtype, "avg", item["avg"], ""])
                ws_an.append([qid, qt, qtype, "max", item["max"], ""])
            for h in item.get("histogram", []):
                ws_an.append([qid, qt, qtype, "hist", f'value={h["value"]}', f'count={h["count"]}'])
        elif qtype in ("date","time"):
            ws_an.append([qid, qt, qtype, "count", item.get("count", 0), ""])
            for v in item.get("by_value", []):
                ws_an.append([qid, qt, qtype, "value", v["value"], f'count={v["count"]}'])
        else:
            ws_an.append([qid, qt, qtype, "count", item.get("count", 0), ""])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"form_{form_id}_export.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )